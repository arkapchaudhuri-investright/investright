"""User analytics — offline geolocation, bot tagging, and sessionisation.

This is the nightly ENRICHMENT layer for the raw `events` table. It never runs
in a web request: `build()` rides the refresh cron, writes the `geo_cache` and
`sessions` caches, and the /admin/analytics dashboard only reads them (the
app's cron-writes / web-reads rule).

What it produces:
  • geo_cache  — every distinct IP → city/region/country + a bot flag, resolved
                 offline against a DB-IP City Lite database (downloaded once to
                 data/; no visitor IP ever leaves the box).
  • sessions   — events grouped into visits (30-min inactivity gap) with a
                 duration, page count, entry/exit path, and funnel milestones.

Time-on-page is the server-side estimate: a page's dwell is the gap to the next
event in the same visit (the last page of a visit has no measurable dwell).
"""
import gzip
import ipaddress
import os
import shutil
import urllib.request
from datetime import datetime, timedelta, timezone

from db import get_conn

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
GEODB_PATH = os.path.join(DATA_DIR, "dbip-city-lite.mmdb")
# DB-IP publishes a free monthly City-Lite mmdb (CC-BY 4.0), no account needed.
GEODB_URL = "https://download.db-ip.com/free/dbip-city-lite-{ym}.mmdb.gz"

SESSION_GAP_S = 30 * 60          # a >30-min silence starts a new visit
DWELL_CAP_S = 30 * 60           # cap one page's counted dwell (idle tabs, lunch)

# Bot heuristics — coarse but effective. UA substrings + a few datacenter/crawler
# IP prefixes we've actually seen. This only *labels*; nothing is dropped.
BOT_UA = ("bot", "crawl", "spider", "slurp", "bingpreview", "facebookexternalhit",
          "headless", "python-requests", "curl", "wget", "httpclient", "scan",
          "monitor", "uptime", "lighthouse", "pagespeed", "semrush", "ahrefs")
BOT_IP_PREFIX = ("66.249.",)   # Googlebot

# Engagement MILESTONES, ordered by depth of intent — deliberately NOT a
# sequential funnel. Search engines land people straight on /stock/<ticker>, so
# "opened a deep-dive" routinely outruns "searched", and a returning account
# holder signs in without saving anything. Treating these as funnel steps
# produced nonsense drop-off figures (−97%), so each is reported independently
# as "% of sessions that ever did this".
FUNNEL = [
    ("visit",    "Landed",              lambda e: True),
    ("deepdive", "Opened a deep-dive",  lambda e: (e["path"] or "").startswith("/stock/")),
    ("search",   "Searched / analyzed", lambda e: e["action"] in ("analyze", "compare")
                                                  or (e["path"] or "").startswith("/analyze")),
    ("signup",   "Signed in",           lambda e: e["user_id"] is not None),
    ("save",     "Saved a stock",       lambda e: e["action"] in ("add", "hold_add", "peer_add",
                                                                   "hold_import")),
]


# ─────────────────────────── geolocation ────────────────────────────
def ensure_geodb():
    """Return a path to the DB-IP City mmdb, downloading it once if absent.
    Tries the current month then the previous two (start-of-month lag). Returns
    None if it can't be fetched — geolocation then degrades to bot-flag only."""
    if os.path.exists(GEODB_PATH):
        return GEODB_PATH
    os.makedirs(DATA_DIR, exist_ok=True)
    now = datetime.now(timezone.utc)
    months = []
    y, m = now.year, now.month
    for _ in range(3):
        months.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    for ym in months:
        try:
            tmp = GEODB_PATH + ".gz"
            # DB-IP 403s the default "Python-urllib/3.x" agent, so the download
            # must identify itself. Without this the fetch failed silently on
            # the VM and every IP resolved with a NULL city/country.
            req = urllib.request.Request(
                GEODB_URL.format(ym=ym),
                headers={"User-Agent": "InvestRight/1.0 (+https://investright.us)"})
            with urllib.request.urlopen(req, timeout=180) as r, open(tmp, "wb") as f:
                shutil.copyfileobj(r, f)
            with gzip.open(tmp, "rb") as fi, open(GEODB_PATH, "wb") as fo:
                shutil.copyfileobj(fi, fo)
            os.remove(tmp)
            return GEODB_PATH
        except Exception:
            continue
    return None


def regeolocate_all(conn):
    """Drop and rebuild every geo_cache row. Needed after the database first
    becomes available: rows resolved while it was missing carry a bot flag but
    a NULL city/country, and geolocate_new() skips IPs it has already seen."""
    conn.execute("DELETE FROM geo_cache")
    return geolocate_new(conn)


def _is_private(ip):
    try:
        return ipaddress.ip_address(ip).is_private
    except ValueError:
        return True                      # unparseable → treat as non-public


def _bot_reason(ip, ua):
    ua = (ua or "").lower()
    if _is_private(ip):
        return "localhost/private"
    if any(ip.startswith(p) for p in BOT_IP_PREFIX):
        return "googlebot"
    if any(tok in ua for tok in BOT_UA):
        return "ua"
    return None


def geolocate_new(conn):
    """Resolve every distinct IP in `events` not yet in geo_cache. Bot flag is
    always set; city/country filled when the mmdb is available. Returns count."""
    try:
        import maxminddb
    except Exception:
        maxminddb = None
    reader = None
    if maxminddb:
        path = ensure_geodb()
        if path:
            try:
                reader = maxminddb.open_database(path)
            except Exception:
                reader = None

    todo = conn.execute(
        "SELECT DISTINCT e.ip, "
        "  (SELECT ua FROM events WHERE ip = e.ip AND ua IS NOT NULL LIMIT 1) ua "
        "FROM events e WHERE e.ip IS NOT NULL AND e.ip != '' "
        "AND e.ip NOT IN (SELECT ip FROM geo_cache)").fetchall()
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    n = 0
    for row in todo:
        ip, ua = row["ip"], row["ua"]
        reason = _bot_reason(ip, ua)
        city = region = country = cc = None
        if reader and not _is_private(ip):
            try:
                rec = reader.get(ip) or {}
                country = (rec.get("country", {}).get("names", {}) or {}).get("en")
                cc = rec.get("country", {}).get("iso_code")
                subs = rec.get("subdivisions") or []
                region = subs[0]["names"].get("en") if subs else None
                city = (rec.get("city", {}).get("names", {}) or {}).get("en")
            except (ValueError, KeyError):
                pass
        conn.execute(
            "INSERT OR REPLACE INTO geo_cache "
            "(ip, city, region, country, country_code, is_bot, bot_reason, resolved_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (ip, city, region, country, cc, 1 if reason else 0, reason, now))
        n += 1
    if reader:
        reader.close()
    return n


# ─────────────────────────── sessionisation ─────────────────────────
def _parse(ts):
    try:
        return datetime.fromisoformat(ts)
    except Exception:
        return None


def rebuild_sessions(conn):
    """Rebuild the `sessions` cache from scratch: order every event by visitor
    then time, cut a new visit on a >30-min gap, and roll up duration, pages,
    entry/exit, funnel milestones, and whether the visit was a bot. Returns the
    number of sessions written."""
    bots = {r["ip"]: r["is_bot"] for r in conn.execute("SELECT ip, is_bot FROM geo_cache")}
    rows = conn.execute(
        "SELECT ts, visitor, name, market, action, path, ua, ip, user_id "
        "FROM events WHERE visitor IS NOT NULL ORDER BY visitor, ts").fetchall()

    conn.execute("DELETE FROM sessions")
    sessions = []
    cur = None

    def flush(s):
        if not s:
            return
        start, end = s["events"][0], s["events"][-1]
        t0, t1 = _parse(start["ts"]), _parse(end["ts"])
        dur = int((t1 - t0).total_seconds()) if (t0 and t1) else 0
        sid = f"{s['visitor']}:{start['ts']}"
        milestones = {k: 0 for k, *_ in FUNNEL}
        for ev in s["events"]:
            for key, _label, pred in FUNNEL:
                if pred(ev):
                    milestones[key] = 1
        # a visit's identity fields = the richest non-null seen in it
        name = next((e["name"] for e in reversed(s["events"]) if e["name"]), None)
        uid = next((e["user_id"] for e in s["events"] if e["user_id"] is not None), None)
        market = next((e["market"] for e in reversed(s["events"]) if e["market"]), None)
        is_bot = 1 if bots.get(start["ip"], 0) or _bot_reason(start["ip"], start["ua"]) else 0
        sessions.append((
            sid, s["visitor"], start["ts"], end["ts"], dur, len(s["events"]),
            start["path"], end["path"], start["ip"], name, market,
            1 if uid is not None else 0, uid, is_bot,
            milestones["search"], milestones["deepdive"], milestones["save"],
            milestones["signup"]))

    for e in rows:
        e = dict(e)
        if cur and e["visitor"] == cur["visitor"]:
            prev_t, this_t = _parse(cur["events"][-1]["ts"]), _parse(e["ts"])
            if prev_t and this_t and (this_t - prev_t).total_seconds() <= SESSION_GAP_S:
                cur["events"].append(e)
                continue
        flush(cur)
        cur = {"visitor": e["visitor"], "events": [e]}
    flush(cur)

    conn.executemany(
        "INSERT INTO sessions (id, visitor, started_at, ended_at, duration_s, pages, "
        "entry_path, exit_path, ip, name, market, signed_in, user_id, is_bot, "
        "hit_search, hit_deepdive, hit_save, hit_signup) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", sessions)
    return len(sessions)


def build():
    """Nightly entry point: geolocate new IPs, then rebuild sessions. Best-effort
    and self-contained; safe to call from refresh.py. Returns a status string."""
    with get_conn() as conn:
        geo = geolocate_new(conn)
        sess = rebuild_sessions(conn)
    return f"geo +{geo} · sessions {sess}"


# ─────────────────────────── dashboard queries ──────────────────────
# Selectable reporting windows. Keys are what the URL carries; the value is a
# day count (None = all time). A whitelist, so no caller-supplied value ever
# reaches the SQL — the cutoff itself is still passed as a bound parameter.
PERIODS = {"7d": ("Last 7 days", 7), "30d": ("Last 30 days", 30),
           "90d": ("Last 90 days", 90), "all": ("All time", None)}
DEFAULT_PERIOD = "30d"


class Window:
    """The reporting range: either a named preset or an explicit from/to pair.
    Bundled into one object so the queries below don't each carry three
    range arguments. `label` is what the page prints above the numbers."""

    def __init__(self, period=DEFAULT_PERIOD, dt_from=None, dt_to=None):
        self.period = period if period in PERIODS else DEFAULT_PERIOD
        self.dt_from, self.dt_to = dt_from, dt_to   # UTC ISO strings or None
        self.custom = bool(dt_from or dt_to)

    def sql(self, col="s.started_at"):
        return _period(self.period, col, self.dt_from, self.dt_to)

    @property
    def label(self):
        if not self.custom:
            return PERIODS[self.period][0].lower()
        fmt = lambda s: (datetime.fromisoformat(s).astimezone()
                         .strftime("%-d %b %Y, %H:%M") if s else None)
        a, b = fmt(self.dt_from), fmt(self.dt_to)
        if a and b:
            return f"{a} → {b}"
        return f"since {a}" if a else f"up to {b}"


def _where_bot(include_bots):
    # Always qualified to the sessions row (s) — geo_cache also has an is_bot,
    # so an unqualified name is ambiguous wherever the two are joined.
    return "" if include_bots else " AND s.is_bot = 0"


def to_utc(local_str):
    """'2026-07-23T14:30' from <input type="datetime-local"> → UTC ISO8601.

    The picker hands back naive wall-clock text in the viewer's own timezone,
    while `events.ts` is stored UTC — comparing the two directly would silently
    shift every boundary by the UTC offset."""
    if not local_str:
        return None
    try:
        dt = datetime.fromisoformat(local_str)
    except (TypeError, ValueError):
        return None
    if dt.tzinfo is None:
        dt = dt.astimezone()             # naive → interpret as this machine's local
    return dt.astimezone(timezone.utc).isoformat(timespec="seconds")


def _period(period, col="s.started_at", dt_from=None, dt_to=None):
    """(sql_fragment, params) restricting `col` to the chosen window. An explicit
    from/to pair (already UTC, via to_utc) overrides the preset; either bound may
    stand alone."""
    if dt_from or dt_to:
        sql, params = "", []
        if dt_from:
            sql += f" AND {col} >= ?"
            params.append(dt_from)
        if dt_to:
            sql += f" AND {col} <= ?"
            params.append(dt_to)
        return sql, params
    days = PERIODS.get(period, PERIODS[DEFAULT_PERIOD])[1]
    if not days:
        return "", []
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat(timespec="seconds")
    return f" AND {col} >= ?", [cutoff]


def overview(conn, include_bots=False, win=None):
    """Headline KPIs over sessions (humans-only unless include_bots)."""
    wb = _where_bot(include_bots)
    wp, pp = (win or Window()).sql()
    row = conn.execute(
        f"SELECT COUNT(*) sessions, COUNT(DISTINCT visitor) visitors, "
        f"  SUM(signed_in) signed_sessions, "
        f"  AVG(duration_s) avg_dur, AVG(pages) avg_pages, "
        f"  SUM(CASE WHEN pages <= 1 THEN 1 ELSE 0 END) bounces "
        f"FROM sessions s WHERE 1=1{wb}{wp}", pp).fetchone()
    # Bot counts follow the same window, so "filtered" reflects what you're seeing.
    totals = conn.execute(
        f"SELECT COUNT(*) total, SUM(is_bot) bots FROM sessions s WHERE 1=1{wp}",
        pp).fetchone()
    d = dict(row)
    d["bot_sessions"] = totals["bots"] or 0
    d["human_sessions"] = (totals["total"] or 0) - d["bot_sessions"]
    d["bounce_rate"] = (100 * d["bounces"] / d["sessions"]) if d["sessions"] else 0
    return d


def funnel(conn, include_bots=False, win=None):
    """Engagement milestones: how many sessions ever reached each one, as a share
    of all sessions. Independent measures, not sequential steps — see FUNNEL for
    why drop-off between them would be meaningless here."""
    wb = _where_bot(include_bots)
    wp, pp = (win or Window()).sql()
    cols = {
        "visit":    "COUNT(*)",
        "search":   "SUM(hit_search)",
        "deepdive": "SUM(hit_deepdive)",
        "save":     "SUM(hit_save)",
        "signup":   "SUM(hit_signup)",
    }
    row = conn.execute(
        f"SELECT {', '.join(f'{expr} AS {k}' for k, expr in cols.items())} "
        f"FROM sessions s WHERE 1=1{wb}{wp}", pp).fetchone()
    top = row["visit"] or 0
    return [{"key": key, "label": label, "count": row[key] or 0,
             "pct_top": (100 * (row[key] or 0) / top) if top else 0}
            for key, label, _ in FUNNEL]


def recent_sessions(conn, include_bots=False, win=None, limit=100):
    """Most recent visits with geo joined, newest first."""
    wb = _where_bot(include_bots)
    wp, pp = (win or Window()).sql()
    rows = conn.execute(
        f"SELECT s.*, g.city, g.region, g.country, g.country_code, g.bot_reason "
        f"FROM sessions s LEFT JOIN geo_cache g ON g.ip = s.ip "
        f"WHERE 1=1{wb}{wp} ORDER BY s.started_at DESC LIMIT ?",
        pp + [limit]).fetchall()
    return [dict(r) for r in rows]


def session_journey(conn, session_id):
    """The ordered page-by-page journey for one visit, with per-page dwell (the
    gap to the next event, capped; last page = None). Returns (meta, steps)."""
    s = conn.execute(
        "SELECT s.*, g.city, g.region, g.country, g.country_code, g.bot_reason "
        "FROM sessions s LEFT JOIN geo_cache g ON g.ip = s.ip WHERE s.id = ?",
        (session_id,)).fetchone()
    if not s:
        return None, []
    s = dict(s)
    evs = conn.execute(
        "SELECT ts, action, path, ticker, ua FROM events "
        "WHERE visitor = ? AND ts >= ? AND ts <= ? ORDER BY ts",
        (s["visitor"], s["started_at"], s["ended_at"])).fetchall()
    steps = []
    evs = [dict(e) for e in evs]
    for i, e in enumerate(evs):
        dwell = None
        if i + 1 < len(evs):
            a, b = _parse(e["ts"]), _parse(evs[i + 1]["ts"])
            if a and b:
                dwell = min(int((b - a).total_seconds()), DWELL_CAP_S)
        e["dwell_s"] = dwell
        steps.append(e)
    return s, steps


def geo_breakdown(conn, include_bots=False, win=None, by="country", limit=40):
    """Distinct-visitor + session counts grouped by country or city."""
    wb = _where_bot(include_bots)
    wp, pp = (win or Window()).sql()
    grp = "g.country" if by == "country" else \
          "COALESCE(g.city,'(unknown)') || ', ' || COALESCE(g.country,'')"
    rows = conn.execute(
        f"SELECT {grp} place, COUNT(DISTINCT s.visitor) visitors, COUNT(*) sessions "
        f"FROM sessions s LEFT JOIN geo_cache g ON g.ip = s.ip "
        f"WHERE 1=1{wb}{wp} GROUP BY place ORDER BY visitors DESC LIMIT ?",
        pp + [limit]).fetchall()
    return [dict(r) for r in rows]


def top_pages(conn, include_bots=False, win=None, limit=20):
    """Most-viewed paths. Reads `events` directly (not sessions), so the window
    filters on e.ts; joins geo_cache by ip to honour the bot filter."""
    wb = " AND COALESCE(g.is_bot,0) = 0" if not include_bots else ""
    wp, pp = (win or Window()).sql("e.ts")
    rows = conn.execute(
        f"SELECT e.path, COUNT(*) views, COUNT(DISTINCT e.visitor) visitors "
        f"FROM events e LEFT JOIN geo_cache g ON g.ip = e.ip "
        f"WHERE e.path IS NOT NULL{wb}{wp} "
        f"GROUP BY e.path ORDER BY views DESC LIMIT ?", pp + [limit]).fetchall()
    return [dict(r) for r in rows]


def daily_series(conn, include_bots=False, win=None):
    """Sessions + unique visitors per calendar day across the window, gap-filled
    so quiet days plot as zero rather than closing the line over them."""
    wb = _where_bot(include_bots)
    wp, pp = (win or Window()).sql()
    rows = conn.execute(
        f"SELECT substr(s.started_at, 1, 10) day, COUNT(*) sessions, "
        f"  COUNT(DISTINCT s.visitor) visitors "
        f"FROM sessions s WHERE 1=1{wb}{wp} GROUP BY day ORDER BY day", pp).fetchall()
    got = {r["day"]: dict(r) for r in rows}
    if not got:
        return []
    first = datetime.fromisoformat(min(got)).date()
    last = datetime.fromisoformat(max(got)).date()
    out, d = [], first
    while d <= last:
        key = d.isoformat()
        out.append(got.get(key, {"day": key, "sessions": 0, "visitors": 0}))
        d += timedelta(days=1)
    return out


# ─────────────────────────── inline SVG charts ──────────────────────
# Built in Python and dropped straight into the page, the same way metrics.py
# draws the deep-dive charts — no JS charting library, nothing loaded remotely.
def _pts(values, w, h, pad):
    """Map a series to (x, y) pixel points inside a padded box."""
    if not values:
        return []
    hi = max(values) or 1
    span = max(len(values) - 1, 1)
    return [(pad + i * (w - 2 * pad) / span,
             h - pad - (v / hi) * (h - 2 * pad)) for i, v in enumerate(values)]


def traffic_chart(series, width=920, height=200, pad=26):
    """Filled area + line of daily sessions, with unique visitors as a second
    line.

    Returns {'svg', 'width', 'height', 'points'} — the points carry each day's
    plotted x/y plus its raw numbers, so the page can hang the same crosshair
    readout on it that the deep-dive price chart uses. {} when there's nothing
    to draw."""
    if len(series) < 2:
        return {}
    sess = [d["sessions"] for d in series]
    vis = [d["visitors"] for d in series]
    hi = max(sess) or 1
    ps = _pts(sess, width, height, pad)
    pv = _pts(vis, width, height, pad)
    line = " ".join(f"{x:.1f},{y:.1f}" for x, y in ps)
    area = (f"{pad},{height - pad} " + line + f" {width - pad},{height - pad}")
    vline = " ".join(f"{x:.1f},{y:.1f}" for x, y in pv)
    # y gridlines at 0 / half / max
    grid = "".join(
        f'<line x1="{pad}" y1="{y:.1f}" x2="{width - pad}" y2="{y:.1f}" '
        f'stroke="var(--hairline)" stroke-width="1"/>'
        f'<text x="{pad - 6}" y="{y + 3:.1f}" text-anchor="end" font-size="10" '
        f'fill="var(--muted)">{v}</text>'
        for v, y in ((hi, pad), (hi // 2, height / 2), (0, height - pad)))
    # date ticks: first, middle, last
    ticks = ""
    for idx in dict.fromkeys((0, len(series) // 2, len(series) - 1)):
        x = ps[idx][0]
        lbl = datetime.fromisoformat(series[idx]["day"]).strftime("%-d %b")
        anchor = "start" if idx == 0 else "end" if idx == len(series) - 1 else "middle"
        ticks += (f'<text x="{x:.1f}" y="{height - 6}" text-anchor="{anchor}" '
                  f'font-size="10" fill="var(--muted)">{lbl}</text>')
    svg = (
        f'<svg viewBox="0 0 {width} {height}" class="an-chart" role="img" '
        f'aria-label="Daily sessions and unique visitors">{grid}'
        f'<polygon points="{area}" fill="var(--accent)" opacity=".14"/>'
        f'<polyline points="{line}" fill="none" stroke="var(--accent)" stroke-width="2" '
        f'stroke-linejoin="round" vector-effect="non-scaling-stroke"/>'
        f'<polyline points="{vline}" fill="none" stroke="var(--muted)" stroke-width="1.5" '
        f'stroke-dasharray="4 3" vector-effect="non-scaling-stroke"/>'
        f'{ticks}</svg>')
    points = [{"x": round(ps[i][0], 1), "y": round(ps[i][1], 1),
               "vy": round(pv[i][1], 1), "day": series[i]["day"],
               "label": datetime.fromisoformat(series[i]["day"]).strftime("%-d %b %Y"),
               "sessions": series[i]["sessions"], "visitors": series[i]["visitors"]}
              for i in range(len(series))]
    return {"svg": svg, "width": width, "height": height, "points": points}
def export_tables(conn, include_bots=False, win=None):
    """The dashboard as plain tabular data: {sheet_name: (headers, rows)}.
    One source for every export format, so CSV, Excel and print never drift."""
    win = win or Window()
    kpis = overview(conn, include_bots, win)
    tables = {}

    tables["Summary"] = (
        ["Metric", "Value"],
        [["Window", win.label],
         ["Bots included", "yes" if include_bots else "no"],
         ["Unique visitors", kpis["visitors"] or 0],
         ["Sessions", kpis["sessions"] or 0],
         ["Signed-in sessions", kpis["signed_sessions"] or 0],
         ["Avg time on site (s)", round(kpis["avg_dur"] or 0)],
         ["Avg pages per visit", round(kpis["avg_pages"] or 0, 2)],
         ["Bounce rate (%)", round(kpis["bounce_rate"], 1)],
         ["Bot sessions in window", kpis["bot_sessions"]],
         ["Generated (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")]])

    tables["Milestones"] = (
        ["Milestone", "Sessions", "% of sessions"],
        [[f["label"], f["count"], round(f["pct_top"], 1)]
         for f in funnel(conn, include_bots, win)])

    tables["Daily"] = (
        ["Day", "Sessions", "Unique visitors"],
        [[d["day"], d["sessions"], d["visitors"]]
         for d in daily_series(conn, include_bots, win)])

    tables["Cities"] = (
        ["Place", "Visitors", "Sessions"],
        [[g["place"] or "(unknown)", g["visitors"], g["sessions"]]
         for g in geo_breakdown(conn, include_bots, win, by="city", limit=500)])

    tables["Pages"] = (
        ["Path", "Views", "Visitors"],
        [[p["path"], p["views"], p["visitors"]]
         for p in top_pages(conn, include_bots, win, limit=500)])

    tables["Visits"] = (
        ["Started", "Ended", "Duration (s)", "Pages", "City", "Country", "IP",
         "Name (self-reported)", "Signed in", "Likely bot", "Entry", "Exit",
         "Deep-dive", "Searched", "Saved"],
        [[s["started_at"], s["ended_at"], s["duration_s"], s["pages"],
          s["city"] or "", s["country"] or "", s["ip"] or "", s["name"] or "",
          "yes" if s["signed_in"] else "", "yes" if s["is_bot"] else "",
          s["entry_path"] or "", s["exit_path"] or "",
          "yes" if s["hit_deepdive"] else "", "yes" if s["hit_search"] else "",
          "yes" if s["hit_save"] else ""]
         for s in recent_sessions(conn, include_bots, win, limit=5000)])
    return tables


def to_xlsx(tables):
    """Multi-sheet .xlsx (bytes) — one sheet per table, bold frozen headers and
    auto-ish column widths. openpyxl is a pure-Python dependency, no system
    libraries, so it installs cleanly on the VM."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, rows) in tables.items():
        ws = wb.create_sheet(name[:31])          # Excel caps sheet names at 31
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, start=1):
            widest = max([len(str(h))] +
                         [len(str(r[i - 1])) for r in rows[:200] if i <= len(r)] or [0])
            ws.column_dimensions[get_column_letter(i)].width = min(max(widest + 2, 10), 44)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(tables):
    """All tables in one CSV, separated by a blank line and a section header —
    Excel and Sheets both open it directly."""
    import csv
    import io

    out = io.StringIO()
    w = csv.writer(out)
    for name, (headers, rows) in tables.items():
        w.writerow([f"# {name}"])
        w.writerow(headers)
        w.writerows(rows)
        w.writerow([])
    return out.getvalue()
